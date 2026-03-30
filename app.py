from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, session
import csv
import os
import io
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = 'projectalloc2024secretkey'

UPLOAD_FOLDER = 'uploads'
PROBLEMS_FILE = os.path.join(UPLOAD_FOLDER, 'problems.json')
REGISTRATIONS_FILE = os.path.join(UPLOAD_FOLDER, 'registrations.json')

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

ADMIN_PASSWORD = "admin123"

def load_problems():
    if os.path.exists(PROBLEMS_FILE):
        with open(PROBLEMS_FILE, 'r') as f:
            return json.load(f)
    return []

def save_problems(problems):
    with open(PROBLEMS_FILE, 'w') as f:
        json.dump(problems, f, indent=2)

def load_registrations():
    if os.path.exists(REGISTRATIONS_FILE):
        with open(REGISTRATIONS_FILE, 'r') as f:
            return json.load(f)
    return {}

def save_registrations(registrations):
    with open(REGISTRATIONS_FILE, 'w') as f:
        json.dump(registrations, f, indent=2)

def get_taken_problems():
    registrations = load_registrations()
    return {v['problem_id'] for v in registrations.values()}

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/admin')
def admin():
    if not session.get('admin'):
        return redirect(url_for('admin_login'))
    problems = load_problems()
    registrations = load_registrations()
    taken = get_taken_problems()
    return render_template('admin.html', problems=problems, registrations=registrations, taken=taken)

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    error = None
    if request.method == 'POST':
        if request.form.get('password') == ADMIN_PASSWORD:
            session['admin'] = True
            return redirect(url_for('admin'))
        error = "Invalid password"
    return render_template('admin_login.html', error=error)

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin', None)
    return redirect(url_for('index'))

@app.route('/admin/upload_csv', methods=['POST'])
def upload_csv():
    if not session.get('admin'):
        return jsonify({'error': 'Unauthorized'}), 401
    
    file = request.files.get('file')
    if not file or not file.filename.endswith('.csv'):
        return jsonify({'error': 'Please upload a valid CSV file'}), 400
    
    problems = []
    stream = io.StringIO(file.stream.read().decode('utf-8'))
    reader = csv.DictReader(stream)
    
    for i, row in enumerate(reader):
        problem = {
            'id': str(i + 1),
            'title': row.get('title', row.get('Title', row.get('Problem Title', row.get('problem_title', '')))).strip(),
            'description': row.get('description', row.get('Description', row.get('Problem Description', row.get('problem_description', '')))).strip(),
            'domain': row.get('domain', row.get('Domain', row.get('Category', row.get('category', 'General')))).strip(),
        }
        if problem['title']:
            problems.append(problem)
    
    save_problems(problems)
    return jsonify({'success': True, 'count': len(problems)})

@app.route('/admin/clear_problems', methods=['POST'])
def clear_problems():
    if not session.get('admin'):
        return jsonify({'error': 'Unauthorized'}), 401
    save_problems([])
    return jsonify({'success': True})

@app.route('/admin/clear_registrations', methods=['POST'])
def clear_registrations():
    if not session.get('admin'):
        return jsonify({'error': 'Unauthorized'}), 401
    save_registrations({})
    return jsonify({'success': True})

@app.route('/admin/download')
def download():
    if not session.get('admin'):
        return redirect(url_for('admin_login'))
    
    fmt = request.args.get('format', 'excel')
    registrations = load_registrations()
    problems = {p['id']: p for p in load_problems()}
    
    rows = []
    for roll_no in sorted(registrations.keys(), key=lambda x: int(x)):
        reg = registrations[roll_no]
        prob = problems.get(reg['problem_id'], {})
        rows.append({
            'Roll No': roll_no,
            'Name': reg['name'],
            'Branch': reg['branch'],
            'Email': reg['email'],
            'Mobile': reg['mobile'],
            'Problem ID': reg['problem_id'],
            'Problem Title': prob.get('title', 'N/A'),
            'Domain': prob.get('domain', 'N/A'),
            'Registered At': reg.get('timestamp', '')
        })
    
    if fmt == 'csv':
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=['Roll No','Name','Branch','Email','Mobile','Problem ID','Problem Title','Domain','Registered At'])
        writer.writeheader()
        writer.writerows(rows)
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode()),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'registrations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        )
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registrations"
        
        headers = ['Roll No','Name','Branch','Email','Mobile','Problem ID','Problem Title','Domain','Registered At']
        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        header_font = Font(color="e94560", bold=True, size=11)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        alt_fill = PatternFill(start_color="f0f0f8", end_color="f0f0f8", fill_type="solid")
        for row_idx, row in enumerate(rows, 2):
            for col, key in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col, value=row[key])
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
        
        col_widths = [10, 22, 15, 30, 15, 12, 40, 18, 22]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        ws.row_dimensions[1].height = 22
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'registrations_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

@app.route('/student')
def student():
    problems = load_problems()
    taken = get_taken_problems()
    registrations = load_registrations()
    return render_template('student.html', problems=problems, taken=taken, registrations=registrations)

@app.route('/api/problems')
def api_problems():
    problems = load_problems()
    taken = get_taken_problems()
    return jsonify({'problems': problems, 'taken': list(taken)})

@app.route('/api/register', methods=['POST'])
def register():
    data = request.json
    roll_no = str(data.get('roll_no', '')).strip()
    name = data.get('name', '').strip()
    branch = data.get('branch', '').strip()
    email = data.get('email', '').strip()
    mobile = data.get('mobile', '').strip()
    problem_id = str(data.get('problem_id', '')).strip()
    
    if not all([roll_no, name, branch, email, mobile, problem_id]):
        return jsonify({'error': 'All fields are required'}), 400
    
    try:
        roll_int = int(roll_no)
        if not (1 <= roll_int <= 60):
            return jsonify({'error': 'Roll number must be between 1 and 60'}), 400
    except:
        return jsonify({'error': 'Invalid roll number'}), 400
    
    registrations = load_registrations()
    
    if roll_no in registrations:
        return jsonify({'error': f'Roll number {roll_no} is already registered'}), 400
    
    taken = get_taken_problems()
    if problem_id in taken:
        return jsonify({'error': 'This problem statement has already been taken!'}), 400
    
    problems = load_problems()
    problem_exists = any(p['id'] == problem_id for p in problems)
    if not problem_exists:
        return jsonify({'error': 'Invalid problem selected'}), 400
    
    registrations[roll_no] = {
        'roll_no': roll_no,
        'name': name,
        'branch': branch,
        'email': email,
        'mobile': mobile,
        'problem_id': problem_id,
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    save_registrations(registrations)
    
    return jsonify({'success': True, 'message': f'Successfully registered! Problem assigned.'})

@app.route('/api/stats')
def stats():
    problems = load_problems()
    registrations = load_registrations()
    taken = get_taken_problems()
    return jsonify({
        'total_problems': len(problems),
        'total_registrations': len(registrations),
        'available': len(problems) - len(taken)
    })

if __name__ == '__main__':
    app.run(debug=True, port=5000)
